from database import database
import re
import numpy as np
import pandas as pd
from email import encoders
import ssl
import smtplib
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import os
from private import Private
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from round_decimal import round_decimal

nbrHawkish = 0
nbrDovish = 0
#
#Param:
#t_name: table name
#desc:description of the table
#isOsc: is it an oscillator
#datas: growth rate of the table
#results: tuple of table
#This function organize the datas into their modules
#return: results
#
def organizeDataPerModule(t_name, desc, isOsc ,datas, results, sourceData):
    money_credit = "^T_(M1SL|M2SL|BOGZ1FL893169105Q|BUSLOANS|TOTALSL)".lower()
    econo = ("^T_(INDPRO|NOCDFSA066MSFRBPHI|PCE|PAYEMS|AWHMAN|USALOLITONOSTSAM|HOUST|PERMIT"
+"|IC4WSA|GACDFSA066MSFRBPHI|HTRUCKSSA|BOGZ1FL145020011Q"
+"|us_leading_index_1968|building_permits_25|chicago_pmi_38|total_vehicle_sales_85"
+"|ism_manufacturing_pmi_173|durable_goods_orders_86|retail_sales_256)").lower()
    inflation = "^T_(CPIAUCSL|PPIACO|AHETPI)".lower()
    fed_pol= "^T_(NFORBRES|BOGNONBR|REQRESNS|T10YFF|DTB3|FEDFUNDS|INTDSRUSM193N)".lower()
    moneyCreditGrowth= "MoneyCreditGrowth"
    economicGrowth="EconomicGrowth"
    inflationTxt = "Inflation"
    fedPolicy= "Fed Policy"
    nameIndicator = t_name.replace("t_","")
    listTmp = list(datas[0])
    listTmp[len(listTmp)-1] = datas[0][len(datas[0])-1].strftime("%Y-%m-%d")

    if listTmp[len(listTmp)-3]:
        global nbrHawkish
        nbrHawkish+=1
    else:
        global nbrDovish
        nbrDovish+=1

    datas[0] = tuple(listTmp)

    if re.search(money_credit,t_name):
        datas[0] = (moneyCreditGrowth,nameIndicator, desc, sourceData)+datas[0]
        results[0].append(datas[0])

    elif re.search(econo,t_name):
        datas[0] = (economicGrowth,nameIndicator, desc, sourceData)+datas[0]
        results[1].append(datas[0])

    elif re.search(inflation,t_name):
        datas[0] = (inflationTxt,nameIndicator, desc, sourceData)+datas[0]
        results[2].append(datas[0])

    elif re.search(fed_pol,t_name):
        datas[0] = (fedPolicy,nameIndicator, desc, sourceData)+datas[0]
        results[3].append(datas[0])
    else:
         print("Not Catagorize: "+t_name)

    return results
#
#Param:
#email_sender: emails of the bot email
#email_receiver: emails of the receiver
#filename: the filename of the growth rate report
#This function send the emails to the email_receiver
#return: results
#
def sendEmail(email_sender,email_receiver,filename,modifyRowIndex,columnNames,indicators):
    pwd = os.getenv('MAIL_BOT_PWD')

    subject =" Growth Rate Report"
    global nbrDovish
    global nbrHawkish
    changeRow =""
    index = 0
    for i in range(len(modifyRowIndex)):

        if(modifyRowIndex[i]):
            index = i 
            fieldRow = ""
            for field in modifyRowIndex[i]:
                field = columnNames[field]
                fieldRow+="    {}\n".format(field)
            changeRow+="Columns of {} changed :\n".format(indicators[index])+fieldRow+"\n    "

    if(len(changeRow) == 0):
        changeRow = "There is no change.\n"

    message = """
    Good morning,\n
    This is the growth rate report.\n
    {}
    Here is the CBStance summary:\n
    number of hawkish:{}\n
    number of dovish:{}\n
    Have a nice day\n
    macroBot.
    """.format(changeRow, nbrHawkish, nbrDovish)

    print(message)

    em = MIMEMultipart()

    em['From'] = email_sender
    em['To'] = email_receiver
    em['Subject'] = subject

    # Open PDF file in binary mode
    with open(filename, "rb") as attachment:
        # Add file as application/octet-stream
        # Email client can usually download this automatically as attachment
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())

    # Encode file in ASCII characters to send by email    
    encoders.encode_base64(part)

    # Add header as key/value pair to attachment part
    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {filename}",
    )

    # Add attachment to message and convert message to string
    em.attach(part)


    em.attach(MIMEText(message, "plain"))

    context = ssl.create_default_context()

    with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
        smtp.login(email_sender, pwd)
        smtp.sendmail(email_sender,email_receiver, em.as_string())
#
#Param:
#columnNames: Columns of the report
#sum_CBStance: array of array containg the sum
#padding_column: extra empty data to fill the column
#results: The data of the report
#This function add the sum field to the report
#return: results
#
def addSummaryCBStance(columnNames, sum_CBStance, padding_column, results):

    index_CBStance = 0

    #Concatenate the data into one dataframe
    for result in results:
        for i in range(len(result)):
            temp_res = list(result[i])
            #Ajoute de somme de cbstance a la 1er et 2ieme range
            if(index_CBStance < len(sum_CBStance)):
                temp_res+= sum_CBStance[index_CBStance]
                result[i] = tuple(temp_res)
                index_CBStance+=1
            #Le reste ajoute du vide pour matcher le nombre de colonne
            else:
                temp_res+= [""] * padding_column
                result[i] = tuple(temp_res)

        my_array = np.array(result)
        df = pd.DataFrame(my_array, columns = columnNames,)
        mainDf.append(df)

    return results
#
#This function tag the modfiy data of each row
#return: results
#
def tagModifyRow(mainDf):
    modifyRowIndex= []
    previousDf = pd.read_csv("./reportGrowthrate.csv").fillna('')
    newDf=pd.concat(mainDf, ignore_index=True)

    #Check if Data same shape in order to highlight real change
    if(previousDf.shape == newDf.shape):

        #Iterate each row previous data or row new data
        for i in range(previousDf.shape[0]):

            #replace None data of oscillator with empty string
            #Else convert everyting to Decimal
            previousRowData = previousDf.iloc[i].values.tolist()
            newRowData = newDf.iloc[i].values.tolist()
            
            if(previousRowData[4] == ''):
                previousRowData[4]= None
            else:
                previousRowData[4] = round_decimal(previousRowData[4])
            previousRowData[5] = round_decimal(previousRowData[5])

            modifyIndex = []
            #tag modify field
            for j in range(len(previousRowData)):
                if not (previousRowData[j] == newRowData[j]):
                    modifyIndex.append(j)
            modifyRowIndex.append(modifyIndex)
            
    return modifyRowIndex
#
#Param:
#fichierCSV: name of the csv file
#fichierExcel: name of the excel file
#This function create the excel with custom
#
def createExcelFileWithHighLight(fichierCSV, fichierExcel):
        
    #Save the csv into excel file
    wb = Workbook()
    ws = wb.active
    #Copy each row
    with open(fichierCSV, 'r') as f:
        for row in csv.reader(f):
            ws.append(row)

    #Yellow
    highlight_y = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    #Orange
    highlight_o = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    #Light blue color 
    highlight_b = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    for cell in ws[16]:  #Highlight row 16
        cell.fill = highlight_b
    for cell in ws[22]:  #Highlight row 22
        cell.fill = highlight_b
        
    # Highlight each wrong row
    for i_row in range(len(modifyRowIndex)):  # Loop through the row numbers
        if(len(modifyRowIndex[i_row]) != 0):

            # Highlight each modify row
            for cell in ws[i_row+2]:  # Loop through each cell in the row
                cell.fill = highlight_y
            # Highlight each modify field of row
            for mod_cell in modifyRowIndex[i_row]:  # Loop through modify cell in the row
                ws[i_row+2][mod_cell].fill = highlight_o
        
    wb.save(fichierExcel)

###debut code###
pr = Private()


db = database(os.getenv('DB_NAME'),os.getenv('DB_USER'),os.getenv('DB_PASSWORD'))
list_links = db.fetch_links("length(show_more)",">","2")
indicators = db.fetch_indicators()
results=([], [], [], [])
#Organise the growth rate for investing
for link in list_links:
    datas = db.fetch_table_show_gr(link[4]+"_gr","t1"
+" inner join " +link[4]+"_gr"+ " t2"
+" on t1.date = t2.date and t2.intervalmonth = 12"
+" join (select tb.t_name , tb.is_oscillator from"
+" (select tl.t_name, tl.is_oscillator from t_links_inv tl"
+" union"
+" select ti.t_name, ti.is_oscillator from t_indicators_fred ti) tb) tb"
+" on tb.t_name = '"+link[4]+"'"
+" where t1.value is not NULL and t1.intervalmonth = 3 and  t2.value is not NULL"
+" order by t1.date desc"
+" limit 1")
    print(link[4].lower())
    results = organizeDataPerModule(link[4].lower(),link[5], link[6] ,datas, results, "investing")

#Organise the growth rate for fred
for indicator in indicators:
    datas = db.fetch_table_show_gr(indicator[3].lower()+"_gr","t1"
+" inner join " +indicator[3].lower()+"_gr"+ " t2"
+" on t1.date = t2.date and t2.intervalmonth = 12"
+" join (select tb.t_name , tb.is_oscillator from"
+" (select tl.t_name, tl.is_oscillator from t_links_inv tl"
+" union"
+" select ti.t_name, ti.is_oscillator from t_indicators_fred ti) tb) tb"
+" on tb.t_name = '"+indicator[3]+"'"
+" where t1.value is not NULL and t1.intervalmonth = 3 and  t2.value is not NULL"
+" order by t1.date desc"
+" limit 1")
    results = organizeDataPerModule(indicator[3].lower(), indicator[1], indicator[4], datas, results, "fred")

mainDf = []
padding_of_sum= 2
sum_CBStance = []
modifyRowIndex = []

#Create the sum hawkish table
empty_string_list = [""] * padding_of_sum
empty_string_list.append("Sum of Hawkish: ")
empty_string_list.append(nbrHawkish)
sum_CBStance.append(empty_string_list)

#Create the sum Dovish table
empty_string_list = [""] * 2
empty_string_list.append("Sum of Dovish: ")
empty_string_list.append(nbrDovish)
sum_CBStance.append(empty_string_list)

padding_column = len(sum_CBStance) * padding_of_sum
indicators = [
    "m1sl",
    "m2sl",
    "bogz1fl893169105q",
    "busloans",
    "totalsl",
    "us_leading_index_1968",
    "building_permits_25",
    "chicago_pmi_38",
    "total_vehicle_sales_85",
    "ism_manufacturing_pmi_173",
    "durable_goods_orders_86",
    "retail_sales_256",
    "indpro",
    "pce",
    "payems",
    "awhman",
    "houst",
    "ic4wsa",
    "bogz1fl145020011q",
    "gacdfsa066msfrbphi",
    "cpiaucsl",
    "ppiaco",
    "ahetpi",
    "t10yff",
    "dtb3",
    "fedfunds"
]

columnNames= ['Module','Indicator','Description',"Source of Data",'3 Month Ann.','12 Month/ 1 Year Growth', 'Has crossover/is Positive', 'CB Stance', 'Date']+[""] * padding_column

results = addSummaryCBStance(columnNames, sum_CBStance, padding_column, results)
currentPath = os.getcwd()
fichierExcel = currentPath+"\\reportGrowthrate.xlsx"
fichierCSV = currentPath+"\\reportGrowthrate.csv"

if os.path.isfile(fichierCSV):
    modifyRowIndex = tagModifyRow(mainDf)

#Check if all index has no change or there is no report in csv or excel, if so dont send email(rechcheck)
#if( not all(rowIndex  for rowIndex in modifyRowIndex) or not (os.path.isfile(fichierCSV)) or not (os.path.isfile(fichierExcel))):

if os.path.isfile(fichierExcel):
    if os.path.isfile(currentPath+"\\previous_reportGrowthrate.xlsx"):
        os.remove(currentPath+"\\previous_reportGrowthrate.xlsx")
    os.rename(fichierExcel,currentPath+"\\previous_reportGrowthrate.xlsx")

if os.path.isfile(fichierCSV):
    os.remove(fichierCSV)

currentPath = os.getcwd()
#Create the csv
pd.concat(mainDf, ignore_index=True).to_csv(currentPath+"\\reportGrowthrate.csv", index=False)
#Create Excel file
createExcelFileWithHighLight(fichierCSV, fichierExcel)

sendEmail(os.getenv('MAIL_BOT'),os.getenv('MAIL_BOT_DEST'), fichierExcel, modifyRowIndex, columnNames, indicators)

db.update_status("process_investing", 0)
db.update_status("process_fred", 0)
pr.clean()